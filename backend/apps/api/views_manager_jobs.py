import logging
import io
from collections import defaultdict
from datetime import datetime, timedelta, time

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core.mail import EmailMessage
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q

from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError

from apps.accounts.models import Company, User
from apps.jobs.models import (
    File,  # оставляем для совместимости
    Job,
    JobCheckEvent,
    JobChecklistItem,
    JobPhoto,
)
from apps.marketing.models import ReportEmailLog
from apps.locations.models import Location

from .pdf import generate_job_report_pdf
from .permissions import IsManagerUser as IsManager
from .serializers import (
    JobChecklistItemSerializer,
    JobCheckEventSerializer,
    JobDetailSerializer,
    ManagerJobCreateSerializer,
    PlanningJobSerializer,
    compute_sla_status_for_job,
    compute_sla_reasons_for_job,
)
logger = logging.getLogger(__name__)

# Console roles that have access to manager endpoints
# Owner = Billing Admin, Manager = Ops Admin, Staff = Limited Access
CONSOLE_ROLES = {User.ROLE_OWNER, User.ROLE_MANAGER, User.ROLE_STAFF}

VALID_SLA_REASONS = {
    "missing_before_photo",
    "missing_after_photo",
    "checklist_not_completed",
    "missing_check_in",
    "missing_check_out",
}


class JobPdfReportView(APIView):
    """
    Генерация PDF отчета по job.

    Доступно:
    - клинеру (по своим job)
    - менеджеру (по job своей компании)
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int):
        user = request.user

        # Allow cleaners (for their own jobs) and console users (for company jobs)
        allowed_roles = {User.ROLE_CLEANER} | CONSOLE_ROLES
        if user.role not in allowed_roles:
            return Response(
                {"detail": "Only cleaners and console users can generate PDF reports."},
                status=status.HTTP_403_FORBIDDEN,
            )

        base_qs = Job.objects.select_related("location", "cleaner").prefetch_related(
            "checklist_items",
            "check_events",
            "photos__file",
        )

        if user.role == User.ROLE_CLEANER:
            job = get_object_or_404(base_qs, pk=pk, cleaner=user)
        else:
            # Console users can access any company job
            job = get_object_or_404(base_qs, pk=pk, company=user.company)

        pdf_bytes = generate_job_report_pdf(job)

        filename = f"job_report_{job.id}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


class ManagerJobPdfEmailView(APIView):
    """
    Реальная отправка PDF-отчёта на email менеджера.

    POST /api/manager/jobs/<id>/report/email/
    Body (опционально): { "email": "manager@example.com" }

    Если email в body не передан — используем email текущего менеджера.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can email PDF reports."},
                status=status.HTTP_403_FORBIDDEN,
            )

        # job только внутри компании менеджера
        job = get_object_or_404(
            Job.objects.filter(company=user.company),
            pk=pk,
        )

        # email можно явно передать в body, иначе берём email текущего юзера
        target_email = (request.data.get("email") or user.email or "").strip()
        if not target_email:
            return Response(
                {"detail": "Email is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # генерируем PDF тем же helper'ом, что и download-эндпоинт
        try:
            pdf_bytes = generate_job_report_pdf(job)
        except Exception:
            return Response(
                {"detail": "Failed to generate PDF report."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        if not pdf_bytes:
            return Response(
                {"detail": "PDF generation returned empty content."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # собираем письмо
        subject = f"Job report #{job.id}"
        location_obj = getattr(job, "location", None)
        location_name = getattr(location_obj, "name", "") or getattr(
            job, "location_name", ""
        ) or ""
        address = getattr(location_obj, "address", "") or getattr(
            job, "location_address", ""
        ) or ""
        scheduled_date = (
            job.scheduled_date.strftime("%Y-%m-%d")
            if getattr(job, "scheduled_date", None)
            else ""
        )
        cleaner_obj = getattr(job, "cleaner", None)
        cleaner_name = getattr(cleaner_obj, "full_name", "") or ""

        sla_status = getattr(job, "sla_status", None)
        sla_reasons = getattr(job, "sla_reasons", None) or []
        if not isinstance(sla_reasons, (list, tuple)):
            sla_reasons = [sla_reasons] if sla_reasons else []

        if sla_status == "violated" and sla_reasons:
            reasons_text = ", ".join(str(r).replace("_", " ") for r in sla_reasons)
            sla_line = f"SLA status: violated (reasons: {reasons_text})"
        elif sla_status:
            sla_line = f"SLA status: {sla_status}"
        else:
            sla_line = ""

        body_lines = [
            "Hello,",
            "",
            "Attached is the verified cleaning report for the completed job.",
            "",
            f"Job ID: {job.id}",
        ]
        if location_name:
            body_lines.append(f"Location: {location_name}")
        if address:
            body_lines.append(f"Address: {address}")
        if scheduled_date:
            body_lines.append(f"Date: {scheduled_date}")
        if cleaner_name:
            body_lines.append(f"Cleaner: {cleaner_name}")
        if sla_line:
            body_lines.append(sla_line)

        body_lines.append("")
        body_lines.append(
            "This report includes check-in/check-out verification, before/after photos, checklist and SLA status."
        )
        body_lines.append("")
        body_lines.append("Regards,")
        body_lines.append("CleanProof")

        body = "\n".join(body_lines)

        from_email = getattr(
            settings,
            "DEFAULT_FROM_EMAIL",
            getattr(settings, "FOUNDER_DEMO_EMAIL", None),
        ) or target_email

        message = EmailMessage(
            subject=subject,
            body=body,
            from_email=from_email,
            to=[target_email],
        )
        message.attach(
            f"job_report_{job.id}.pdf",
            pdf_bytes,
            "application/pdf",
        )

        try:
            # отправляем письмо
            message.send(fail_silently=False)

            # ✅ логируем успешную отправку
            try:
                ReportEmailLog.objects.create(
                    company_id=job.company_id,
                    user=user,
                    kind=ReportEmailLog.KIND_JOB_REPORT,
                    job_id=job.id,
                    period_from=None,
                    period_to=None,
                    to_email=target_email,
                    subject=subject,
                    status=ReportEmailLog.STATUS_SENT,
                    error_message="",
                )
            except Exception as log_exc:
                logger.exception(
                    "Failed to log sent job report email", exc_info=log_exc
                )

        except Exception as exc:
            # ❌ логируем неудачную попытку
            try:
                ReportEmailLog.objects.create(
                    company_id=job.company_id,
                    user=user,
                    kind=ReportEmailLog.KIND_JOB_REPORT,
                    job_id=job.id,
                    period_from=None,
                    period_to=None,
                    to_email=target_email,
                    subject=subject,
                    status=ReportEmailLog.STATUS_FAILED,
                    error_message=str(exc),
                )
            except Exception as log_exc:
                logger.exception(
                    "Failed to log failed job report email",
                    exc_info=log_exc,
                )

            return Response(
                {"detail": "Failed to send email."},
                status=status.HTTP_502_BAD_GATEWAY,
            )

        return Response(
            {
                "detail": "PDF report emailed.",
                "job_id": job.id,
                "target_email": target_email,
            },
            status=status.HTTP_200_OK,
        )


class ManagerJobReportEmailLogListView(APIView):
    """
    История отправок PDF-отчёта по конкретной job.

    GET /api/manager/jobs/<id>/report/emails/
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can view report email history."},
                status=status.HTTP_403_FORBIDDEN,
            )

        job = get_object_or_404(
            Job.objects.filter(company=user.company),
            pk=pk,
        )

        logs = (
            ReportEmailLog.objects.filter(
                company_id=user.company_id,
                kind=ReportEmailLog.KIND_JOB_REPORT,
                job_id=job.id,
            )
            .select_related("user")
            .order_by("-created_at")[:20]
        )

        payload = {
            "job_id": job.id,
            "emails": [
                {
                    "id": log.id,
                    "sent_at": log.created_at.isoformat(),
                    "target_email": log.to_email,
                    "status": log.status,
                    "sent_by": (
                        getattr(log.user, "full_name", None)
                        or getattr(log.user, "email", None)
                        if log.user
                        else None
                    ),
                    "subject": log.subject,
                    "error_message": log.error_message,
                }
                for log in logs
            ],
        }

        return Response(payload, status=status.HTTP_200_OK)


class ManagerJobsTodayView(APIView):
    """
    Today jobs для менеджера (dashboard).

    GET /api/manager/jobs/today/
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can view jobs overview."},
                status=status.HTTP_403_FORBIDDEN,
            )

        today = timezone.localdate()

        qs = (
            Job.objects.filter(company=user.company, scheduled_date=today)
            .select_related("location", "cleaner")
            .order_by("scheduled_start_time", "id")
        )

        data = []
        for job in qs:
            location = job.location
            cleaner = job.cleaner
            data.append(
                {
                    "id": job.id,
                    "status": job.status,
                    "scheduled_date": job.scheduled_date,
                    "scheduled_start_time": job.scheduled_start_time,
                    "scheduled_end_time": job.scheduled_end_time,
                    "location": {
                        "id": getattr(location, "id", None),
                        "name": getattr(location, "name", None),
                        "address": getattr(location, "address", None),
                    },
                    "cleaner": {
                        "id": getattr(cleaner, "id", None),
                        "full_name": getattr(cleaner, "full_name", None),
                        "phone": getattr(cleaner, "phone", None),
                    },
                }
            )

        return Response(data, status=status.HTTP_200_OK)


# Сколько дней показывать в Completed на странице Jobs
ACTIVE_COMPLETED_DAYS = 30


class ManagerJobsActiveView(APIView):
    """
    GET /api/manager/jobs/active/

    Оперативный список джоб для вкладок Jobs:
    - все НЕ completed (любая дата);
    - completed за последние ACTIVE_COMPLETED_DAYS дней.

    Остальной полный архив — через Job History.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, IsManager]

    def get(self, request):
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"detail": "Manager has no company."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = timezone.localdate()
        completed_from = today - timedelta(days=ACTIVE_COMPLETED_DAYS)

        qs = (
            Job.objects.filter(company=company)
            .filter(
                Q(status__in=[Job.STATUS_SCHEDULED, Job.STATUS_IN_PROGRESS])
                | Q(
                    status=Job.STATUS_COMPLETED,
                    actual_end_time__isnull=False,
                    actual_end_time__date__gte=completed_from,
                )
            )
            .select_related("location", "cleaner")
            .prefetch_related("photos")
            .order_by("scheduled_date", "scheduled_start_time", "id")
        )

        data: list[dict] = []

        for job in qs:
            location = getattr(job, "location", None)
            cleaner = getattr(job, "cleaner", None)

            photos_qs = getattr(job, "photos", None)
            photos = list(photos_qs.all()) if photos_qs is not None else []

            has_before = any(
                p.photo_type == JobPhoto.TYPE_BEFORE for p in photos
            )
            has_after = any(
                p.photo_type == JobPhoto.TYPE_AFTER for p in photos
            )

            data.append(
                {
                    "id": job.id,
                    "status": job.status,
                    "scheduled_date": job.scheduled_date.isoformat()
                    if job.scheduled_date
                    else None,
                    "scheduled_start_time": job.scheduled_start_time.strftime(
                        "%H:%M"
                    )
                    if job.scheduled_start_time
                    else None,
                    "scheduled_end_time": job.scheduled_end_time.strftime(
                        "%H:%M"
                    )
                    if job.scheduled_end_time
                    else None,
                    "location_name": getattr(location, "name", "") or "",
                    "location_address": getattr(location, "address", "") or "",
                    "cleaner_name": getattr(cleaner, "full_name", "") or "",
                    # флаги для has_proof на фронте
                    "has_before_photo": has_before,
                    "has_after_photo": has_after,
                }
            )

        return Response(data, status=status.HTTP_200_OK)


class ManagerJobsCreateView(APIView):
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can create jobs."},
                status=status.HTTP_403_FORBIDDEN,
            )

        company = user.company

        # ⛔ Компания заблокирована (истёк trial или явно заблокирована)
        if company.is_blocked():
            code = "trial_expired" if company.is_trial_expired() else "company_blocked"
            detail = (
                "Your free trial has ended. You can still view existing jobs and "
                "download reports, but creating new jobs requires an upgrade."
                if code == "trial_expired"
                else "Your account is currently blocked. Please contact support."
            )
            return Response(
                {"code": code, "detail": detail},
                status=status.HTTP_403_FORBIDDEN,
            )

        # ⛔ Trial-лимит по количеству jobs
        if company.is_trial_active and company.trial_jobs_limit_reached():
            return Response(
                {
                    "code": "trial_jobs_limit_reached",
                    "detail": (
                        "Your free trial allows up to "
                        f"{Company.TRIAL_MAX_JOBS} jobs. "
                        "Please upgrade your plan to create more jobs."
                    ),
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = ManagerJobCreateSerializer(
            data=request.data,
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)

        # 🚦 Guard: нельзя создавать job на неактивную локацию
        location = serializer.validated_data.get("location")
        if isinstance(location, Location) and not getattr(location, "is_active", True):
            return Response(
                {
                    "code": "location_inactive",
                    "detail": "This location is inactive. Please choose another location.",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        job = serializer.save()
        out = PlanningJobSerializer(job).data
        return Response(out, status=status.HTTP_201_CREATED)


class ManagerJobDetailView(APIView):
    """
    Детали job для менеджера + фото, чеклист, события.

    GET /api/manager/jobs/<id>/
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, pk: int):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can view job details."},
                status=status.HTTP_403_FORBIDDEN,
            )

        job = get_object_or_404(
            Job.objects.select_related("location", "cleaner").prefetch_related(
                "checklist_items",
                "check_events",
                "photos__file",
            ),
            pk=pk,
            company=user.company,
        )

        location = job.location
        cleaner = job.cleaner

        photos_data = []
        for p in job.photos.all().select_related("file").order_by("photo_type", "id"):
            file_url = p.file.file_url if p.file else None
            if file_url and file_url.startswith("/"):
                file_url = request.build_absolute_uri(file_url)

            photos_data.append(
                {
                    "photo_type": p.photo_type,
                    "file_url": file_url,
                    "latitude": p.latitude,
                    "longitude": p.longitude,
                    "photo_timestamp": p.photo_timestamp,
                    "created_at": p.created_at,
                }
            )

        checklist_data = JobChecklistItemSerializer(
            job.checklist_items.all(), many=True
        ).data
        events_data = JobCheckEventSerializer(
            job.check_events.all().order_by("created_at"),
            many=True,
        ).data

        data = {
            "id": job.id,
            "status": job.status,
            "scheduled_date": job.scheduled_date,
            "scheduled_start_time": job.scheduled_start_time,
            "scheduled_end_time": job.scheduled_end_time,
            "actual_start_time": job.actual_start_time,
            "actual_end_time": job.actual_end_time,
            "location": {
                "id": getattr(location, "id", None),
                "name": getattr(location, "name", None),
                "address": getattr(location, "address", None),
                "latitude": getattr(location, "latitude", None),
                "longitude": getattr(location, "longitude", None),
            },
            "cleaner": {
                "id": getattr(cleaner, "id", None),
                "full_name": getattr(cleaner, "full_name", None),
                "phone": getattr(cleaner, "phone", None),
            },
            "manager_notes": job.manager_notes,
            "cleaner_notes": job.cleaner_notes,
            "photos": photos_data,
            "checklist_items": checklist_data,
            "check_events": events_data,
        }

        # SLA-слой: статус и причины нарушений для этой job
        data["sla_status"] = compute_sla_status_for_job(job)
        data["sla_reasons"] = compute_sla_reasons_for_job(job)

        return Response(data, status=status.HTTP_200_OK)


def build_planning_job_payload(job: Job):
    """
    Helper: единый payload для planning/history (один и тот же формат).
    """
    location = job.location
    cleaner = job.cleaner

    # proof: photos
    before_uploaded = False
    after_uploaded = False
    try:
        photos = list(job.photos.all())
    except Exception:
        photos = []

    for p in photos:
        if p.photo_type == JobPhoto.TYPE_BEFORE:
            before_uploaded = True
        elif p.photo_type == JobPhoto.TYPE_AFTER:
            after_uploaded = True

    # proof: checklist (required)
    checklist_completed = False
    try:
        items = list(job.checklist_items.all())
    except Exception:
        items = []

    checklist_items_texts = [
        getattr(it, "text", "").strip()
        for it in items
        if getattr(it, "text", "").strip()
    ]

    if not items:
        checklist_completed = False
    else:
        required_attr = None
        sample = items[0]
        if hasattr(sample, "required"):
            required_attr = "required"
        elif hasattr(sample, "is_required"):
            required_attr = "is_required"

        if required_attr:
            required_items = [
                it for it in items if bool(getattr(it, required_attr, False))
            ]
            if not required_items:
                required_items = items
        else:
            required_items = items

        checklist_completed = all(
            bool(getattr(it, "is_completed", False)) for it in required_items
        )

    sla_reasons: list[str] = []

    if job.status == Job.STATUS_COMPLETED:
        if not before_uploaded:
            sla_reasons.append("missing_before_photo")
        if not after_uploaded:
            sla_reasons.append("missing_after_photo")
        if not checklist_completed:
            sla_reasons.append("checklist_not_completed")

    sla_status = "violated" if sla_reasons else "ok"

    checklist_template = getattr(job, "checklist_template", None)
    checklist_template_payload = None
    if checklist_template is not None:
        checklist_template_payload = {
            "id": checklist_template.id,
            "name": checklist_template.name,
        }

    return {
        "id": job.id,
        "scheduled_date": job.scheduled_date,
        "scheduled_start_time": job.scheduled_start_time,
        "scheduled_end_time": job.scheduled_end_time,
        "status": job.status,
        "location": {
            "id": getattr(location, "id", None),
            "name": getattr(location, "name", None),
            "address": getattr(location, "address", None),
        },
        "cleaner": {
            "id": getattr(cleaner, "id", None),
            "full_name": getattr(cleaner, "full_name", None),
        },
        "proof": {
            "before_uploaded": bool(before_uploaded),
            "after_uploaded": bool(after_uploaded),
            "checklist_completed": bool(checklist_completed),
            "before_photo": bool(before_uploaded),
            "after_photo": bool(after_uploaded),
            "checklist": bool(checklist_completed),
        },
        "sla_status": sla_status,
        "sla_reasons": sla_reasons,
        "checklist_template": checklist_template_payload,
        "checklist_items": checklist_items_texts,
    }


class ManagerJobForceCompleteView(APIView):
    """
    Force-complete job (manager override).

    POST /api/manager/jobs/<id>/force-complete/

    AUDIT FIXES:
    - Critical Risk #1: Block force-complete from scheduled (must check in first)
    - Critical Risk #3: Transitions to completed_unverified (not completed)
    - Critical Risk #4: Persist audit fields (verification_override, force_completed_at, etc.)

    Hybrid Verified Model:
    - Only allowed when job is in_progress (cleaner already checked in with GPS)
    - Transitions to completed_unverified (separated from verified completions)
    - Requires reason text
    - All audit metadata persisted
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, pk: int):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can force-complete jobs."},
                status=status.HTTP_403_FORBIDDEN,
            )

        company = getattr(user, "company", None)
        if company is None:
            return Response(
                {"detail": "User has no company."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if company.is_blocked():
            code = "trial_expired" if company.is_trial_expired() else "company_blocked"
            if code == "trial_expired":
                detail = (
                    "Your free trial has ended. You can still view existing jobs and "
                    "download reports, but overriding jobs requires an upgrade."
                )
            else:
                detail = "Your account is currently blocked. Please contact support."

            return Response(
                {"code": code, "detail": detail},
                status=status.HTTP_403_FORBIDDEN,
            )

        # AUDIT FIX: Critical Risk #2 - Atomic transaction + row-level locking
        with transaction.atomic():
            job = get_object_or_404(
                Job.objects.select_related("location", "cleaner").select_for_update(),
                pk=pk,
                company=company,
            )

            # AUDIT FIX: Block if already completed or completed_unverified
            if job.status in (Job.STATUS_COMPLETED, Job.STATUS_COMPLETED_UNVERIFIED):
                return Response(
                    {"detail": "Job is already completed and cannot be force-completed."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # AUDIT FIX: Critical Risk #1 - Must be in_progress (check-in already happened)
            if job.status != Job.STATUS_IN_PROGRESS:
                return Response(
                    {
                        "detail": (
                            "Force-complete is only allowed for jobs in progress. "
                            "The cleaner must check in first to establish GPS proof. "
                            f"Current status: {job.status}"
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate reason
            reason = (request.data.get("reason") or "").strip()
            if not reason:
                return Response(
                    {"detail": "Reason is required (text explanation for force-complete)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            now = timezone.now()

            # AUDIT FIX: Critical Risk #3 - Transition to completed_unverified
            job.status = Job.STATUS_COMPLETED_UNVERIFIED

            # Set actual_end_time if not already set
            if not job.actual_end_time:
                job.actual_end_time = now

            # AUDIT FIX: Critical Risk #4 - Persist audit fields (now exist on model)
            job.verification_override = True
            job.force_completed_at = now
            job.force_completed_by = user
            job.force_complete_reason = reason

            job.save(update_fields=[
                "status",
                "actual_end_time",
                "verification_override",
                "force_completed_at",
                "force_completed_by_id",
                "force_complete_reason",
            ])

            # Create audit event (TYPE_FORCE_COMPLETE now exists in model)
            JobCheckEvent.objects.create(
                job=job,
                user=user,
                event_type=JobCheckEvent.TYPE_FORCE_COMPLETE,
                # No GPS coordinates (this is the override point)
            )

        response_data = {
            "id": job.id,
            "status": job.status,
            "verification_override": job.verification_override,
            "force_completed_at": job.force_completed_at.isoformat(),
            "force_completed_by": {
                "id": user.id,
                "full_name": getattr(user, "full_name", "") or getattr(user, "email", ""),
            },
            "force_complete_reason": job.force_complete_reason,
        }

        return Response(response_data, status=status.HTTP_200_OK)


class ManagerPlanningJobsView(APIView):
    """
    Job Planning list для менеджера (read-only).

    GET /api/manager/jobs/planning/?date=YYYY-MM-DD
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if user.role not in CONSOLE_ROLES:
            return Response(
                {"detail": "Only console users can view planning jobs."},
                status=status.HTTP_403_FORBIDDEN,
            )

        date_str = (request.query_params.get("date") or "").strip()
        if not date_str:
            return Response(
                {"detail": "date query param is required: YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        day = parse_date(date_str)

        if not day:
            try:
                day = timezone.datetime.strptime(date_str, "%d.%m.%Y").date()
            except Exception:
                day = None

        if not day and "T" in date_str:
            day = parse_date(date_str.split("T", 1)[0])

        if not day:
            return Response(
                {
                    "detail": "Invalid date format. Expected YYYY-MM-DD or DD.MM.YYYY"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = (
            Job.objects.filter(company=user.company, scheduled_date=day)
            .select_related("location", "cleaner")
            .prefetch_related("photos", "checklist_items")
            .order_by("scheduled_start_time", "id")
        )

        data = [build_planning_job_payload(job) for job in qs]
        return Response(data, status=status.HTTP_200_OK)


class ManagerJobsHistoryView(APIView):
    """
    Job History list для менеджера (read-only).

    GET /api/manager/jobs/history/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        user = request.user

        if getattr(user, "role", None) not in (User.ROLE_MANAGER, "manager"):
            return Response(
                {"detail": "Only managers can access job history."},
                status=status.HTTP_403_FORBIDDEN,
            )

        date_from_str = request.query_params.get("date_from")
        date_to_str = request.query_params.get("date_to")

        if not date_from_str or not date_to_str:
            return Response(
                {
                    "detail": "date_from and date_to are required in format YYYY-MM-DD."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"detail": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = (
            Job.objects.filter(
                company=user.company,
                scheduled_date__gte=date_from,
                scheduled_date__lte=date_to,
            )
            .select_related("location", "cleaner")
            .prefetch_related("photos", "checklist_items")
            .order_by("-scheduled_date", "-scheduled_start_time", "-id")
        )

        status_param = request.query_params.get("status")
        if status_param:
            qs = qs.filter(status=status_param)

        cleaner_id = request.query_params.get("cleaner_id")
        if cleaner_id:
            qs = qs.filter(cleaner_id=cleaner_id)

        location_id = request.query_params.get("location_id")
        if location_id:
            qs = qs.filter(location_id=location_id)

        data = [build_planning_job_payload(job) for job in qs]
        return Response(data, status=status.HTTP_200_OK)

class ManagerJobsExportView(APIView):
    """
    XLSX export of completed jobs for managers.

    GET /api/manager/jobs/export/?from=YYYY-MM-DD&to=YYYY-MM-DD[&location_id=&cleaner_id=&sla_status=]

    Exports both human-readable data (names, formatted dates) and technical IDs
    for integration with other systems.
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated, IsManager]

    # SLA reason codes → human-readable labels
    SLA_REASON_LABELS = {
        "missing_before_photo": "Missing Before Photo",
        "missing_after_photo": "Missing After Photo",
        "checklist_not_completed": "Checklist Incomplete",
        "missing_check_in": "Missing Check-in",
        "missing_check_out": "Missing Check-out",
    }

    def get(self, request, *args, **kwargs):
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"detail": "Manager has no company."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1. Validate dates
        date_from_str = request.query_params.get("from")
        date_to_str = request.query_params.get("to")

        if not date_from_str or not date_to_str:
            raise ValidationError(
                {"detail": "`from` and `to` are required (YYYY-MM-DD)."}
            )

        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError(
                {"detail": "Invalid date format. Use YYYY-MM-DD."}
            )

        if date_from > date_to:
            raise ValidationError({"detail": "`from` must be <= `to`."})

        # 2. Date range → aware datetime
        start_dt = timezone.make_aware(datetime.combine(date_from, time.min))
        end_dt = timezone.make_aware(datetime.combine(date_to, time.max))

        # 3. Base queryset: only completed jobs for this company
        qs = (
            Job.objects.select_related("company", "location", "cleaner")
            .filter(
                company=company,
                status=Job.STATUS_COMPLETED,
                actual_end_time__gte=start_dt,
                actual_end_time__lte=end_dt,
            )
            .order_by("-actual_end_time")
        )

        # 4. Additional filters
        location_id = request.query_params.get("location_id")
        if location_id:
            qs = qs.filter(location_id=location_id)

        cleaner_id = request.query_params.get("cleaner_id")
        if cleaner_id:
            qs = qs.filter(cleaner_id=cleaner_id)

        sla_status_filter = request.query_params.get("sla_status")

        # 5. Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job History"

        # 6. Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        violated_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 7. Headers - human-readable + technical IDs
        headers = [
            "Job #",
            "Date",
            "Start Time",
            "End Time",
            "Location",
            "Address",
            "Cleaner",
            "Duration (min)",
            "SLA Status",
            "SLA Issues",
            "Override",
            # Technical columns for integrations
            "Job ID",
            "Location ID",
            "Cleaner ID",
            "Actual Start",
            "Actual End",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 8. Data rows
        row_num = 2
        for job in qs:
            # Compute SLA status dynamically
            computed_sla_status = compute_sla_status_for_job(job)
            computed_sla_reasons = compute_sla_reasons_for_job(job)

            # Filter by sla_status if requested
            if sla_status_filter in ("ok", "violated"):
                if computed_sla_status != sla_status_filter:
                    continue

            # Calculate duration
            duration_minutes = None
            if job.actual_start_time and job.actual_end_time:
                delta = job.actual_end_time - job.actual_start_time
                duration_minutes = int(delta.total_seconds() // 60)

            # Format SLA reasons as human-readable text
            sla_issues_text = ""
            if computed_sla_reasons:
                if isinstance(computed_sla_reasons, (list, tuple)):
                    labels = [
                        self.SLA_REASON_LABELS.get(r, r.replace("_", " ").title())
                        for r in computed_sla_reasons
                    ]
                    sla_issues_text = ", ".join(labels)
                else:
                    sla_issues_text = self.SLA_REASON_LABELS.get(
                        computed_sla_reasons,
                        str(computed_sla_reasons).replace("_", " ").title()
                    )

            # Location and cleaner info
            location = job.location
            cleaner = job.cleaner

            location_name = getattr(location, "name", "") or ""
            location_address = getattr(location, "address", "") or ""
            cleaner_name = getattr(cleaner, "full_name", "") or ""

            # Format times
            scheduled_date = job.scheduled_date.strftime("%Y-%m-%d") if job.scheduled_date else ""
            start_time = job.scheduled_start_time.strftime("%H:%M") if job.scheduled_start_time else ""
            end_time = job.scheduled_end_time.strftime("%H:%M") if job.scheduled_end_time else ""

            # SLA status display
            sla_display = "OK" if computed_sla_status == "ok" else "Issues Found"
            override_display = "Yes" if getattr(job, "verification_override", False) else "No"

            # Row data
            row_data = [
                f"JOB-{job.id:03d}",
                scheduled_date,
                start_time,
                end_time,
                location_name,
                location_address,
                cleaner_name,
                duration_minutes,
                sla_display,
                sla_issues_text,
                override_display,
                # Technical columns
                job.id,
                getattr(job, "location_id", ""),
                getattr(job, "cleaner_id", ""),
                job.actual_start_time.isoformat() if job.actual_start_time else "",
                job.actual_end_time.isoformat() if job.actual_end_time else "",
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

                # Color SLA status column
                if col_num == 9:  # SLA Status column
                    if computed_sla_status == "ok":
                        cell.fill = ok_fill
                    else:
                        cell.fill = violated_fill

            row_num += 1

        # 9. Auto-adjust column widths
        column_widths = [12, 12, 10, 10, 25, 30, 20, 14, 14, 35, 10, 10, 12, 12, 22, 22]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 10. Freeze header row
        ws.freeze_panes = "A2"

        # 11. Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # 12. Return response
        filename = f"jobs_export_{date_from_str}_{date_to_str}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response
